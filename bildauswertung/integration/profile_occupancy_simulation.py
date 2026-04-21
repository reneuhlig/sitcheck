from __future__ import annotations

from collections import defaultdict, deque
from datetime import datetime
import os
import random
import statistics
import threading
import time
from typing import Any, Deque, Dict, List, Optional, Tuple


class ProfileOccupancySimulation:
    """Background occupancy simulator based on historical profile buckets."""

    def __init__(
        self,
        *,
        excel_path: str,
        enabled: bool = False,
        tick_seconds: float = 60.0,
        profile_blend: float = 0.72,
        noise_sigma_scale: float = 0.85,
        max_step_per_tick: float = 2.0,
        rollback_minutes: float = 15.0,
    ):
        self._lock = threading.Lock()
        self._running = False
        self._thread: Optional[threading.Thread] = None

        self.excel_path = str(excel_path)
        self.enabled = bool(enabled)
        # Entry/exit deltas from the profile are emitted at most once per minute.
        self.tick_seconds = max(60.0, float(tick_seconds))
        self.profile_blend = max(0.05, min(1.0, float(profile_blend)))
        self.noise_sigma_scale = max(0.0, min(4.0, float(noise_sigma_scale)))
        self.max_step_per_tick = max(0.2, float(max_step_per_tick))
        self.rollback_minutes = max(1.0, min(240.0, float(rollback_minutes)))

        self._profile_loaded = False
        self._profile_error = ""
        self._profile_bucket_count = 0
        self._profile_by_weekday_slot: Dict[Tuple[int, int], Tuple[float, float]] = {}
        self._profile_by_slot: Dict[int, Tuple[float, float]] = {}
        self._global_mean = 0.0
        self._max_reasonable = 120.0

        self._base_occupancy = 0.0
        self._base_occupancy_rounded = 0
        self._effective_occupancy = 0
        self._tick_entries_buffer = 0
        self._tick_exits_buffer = 0

        self._sim_entries_total = 0
        self._sim_exits_total = 0
        self._detection_impacts: Deque[Dict[str, float]] = deque(maxlen=5000)

        self._last_tick_mono = 0.0
        self._last_profile_timestamp = ""

    def start(self):
        with self._lock:
            if self._running:
                return
            self._running = True
            self._last_tick_mono = time.monotonic()
            self._thread = threading.Thread(target=self._run_loop, daemon=True)
            self._thread.start()

    def stop(self):
        with self._lock:
            self._running = False
            thread = self._thread
        if thread is not None:
            thread.join(timeout=2.0)

    def update_settings(self, payload: Dict[str, Any]):
        payload = dict(payload or {})

        with self._lock:
            if "enabled" in payload:
                self.enabled = bool(payload.get("enabled", False))
            if "excel_path" in payload:
                self.excel_path = str(payload.get("excel_path", self.excel_path)).strip()
                self._profile_loaded = False
                self._profile_error = ""
            if "tick_seconds" in payload:
                self.tick_seconds = max(60.0, float(payload.get("tick_seconds", self.tick_seconds)))
            if "profile_blend" in payload:
                self.profile_blend = max(0.05, min(1.0, float(payload.get("profile_blend", self.profile_blend))))
            if "noise_sigma_scale" in payload:
                self.noise_sigma_scale = max(0.0, min(4.0, float(payload.get("noise_sigma_scale", self.noise_sigma_scale))))
            if "max_step_per_tick" in payload:
                self.max_step_per_tick = max(0.2, float(payload.get("max_step_per_tick", self.max_step_per_tick)))
            if "rollback_minutes" in payload:
                self.rollback_minutes = max(1.0, min(240.0, float(payload.get("rollback_minutes", self.rollback_minutes))))

    def register_detection_event(self, event_type: str):
        normalized = str(event_type or "").strip().lower()
        if normalized not in {"entry", "exit"}:
            return

        delta = 1.0 if normalized == "entry" else -1.0
        with self._lock:
            self._detection_impacts.append({"delta": delta, "ts": time.monotonic()})

    def consume_tick_events(self) -> Dict[str, int]:
        with self._lock:
            values = {
                "entries": int(self._tick_entries_buffer),
                "exits": int(self._tick_exits_buffer),
                "sim_entries_total": int(self._sim_entries_total),
                "sim_exits_total": int(self._sim_exits_total),
            }
            self._tick_entries_buffer = 0
            self._tick_exits_buffer = 0
            return values

    def current_occupancy(self) -> int:
        with self._lock:
            return int(self._effective_occupancy)

    def is_enabled(self) -> bool:
        with self._lock:
            return bool(self.enabled)

    def get_status(self) -> Dict[str, Any]:
        with self._lock:
            return {
                "enabled": bool(self.enabled),
                "excel_path": self.excel_path,
                "profile_loaded": bool(self._profile_loaded),
                "profile_error": self._profile_error,
                "profile_bucket_count": int(self._profile_bucket_count),
                "tick_seconds": float(self.tick_seconds),
                "profile_blend": float(self.profile_blend),
                "noise_sigma_scale": float(self.noise_sigma_scale),
                "max_step_per_tick": float(self.max_step_per_tick),
                "rollback_minutes": float(self.rollback_minutes),
                "simulated_occupancy": int(self._effective_occupancy),
                "sim_entries_total": int(self._sim_entries_total),
                "sim_exits_total": int(self._sim_exits_total),
                "pending_detection_impacts": int(len(self._detection_impacts)),
                "last_profile_timestamp": self._last_profile_timestamp,
            }

    def _run_loop(self):
        while True:
            with self._lock:
                running = bool(self._running)
                enabled = bool(self.enabled)
                tick_seconds = float(self.tick_seconds)
            if not running:
                return

            if enabled:
                try:
                    self._tick()
                except Exception as exc:
                    with self._lock:
                        self._profile_error = f"simulation_tick_failed: {exc}"

            time.sleep(max(0.2, tick_seconds * 0.2))

    def _tick(self):
        now_mono = time.monotonic()
        do_step = False
        with self._lock:
            if not self._profile_loaded:
                self._load_profile_locked()
            if not self._profile_loaded:
                return
            elapsed = max(0.01, now_mono - self._last_tick_mono) if self._last_tick_mono > 0 else self.tick_seconds
            if (now_mono - self._last_tick_mono) >= self.tick_seconds:
                self._last_tick_mono = now_mono
                do_step = True

        if not do_step:
            return

        now_dt = datetime.now()
        target_mean, target_std = self._target_for_datetime(now_dt)

        with self._lock:
            blend = self.profile_blend
            noise_scale = self.noise_sigma_scale
            max_step = self.max_step_per_tick * max(0.5, elapsed / max(0.5, self.tick_seconds))
            rollback_sec = self.rollback_minutes * 60.0

            noisy_target = target_mean + random.gauss(0.0, max(0.5, target_std) * noise_scale)
            noisy_target = max(0.0, min(self._max_reasonable, noisy_target))

            mixed_target = (blend * noisy_target) + ((1.0 - blend) * self._base_occupancy)
            delta = mixed_target - self._base_occupancy
            if delta > max_step:
                delta = max_step
            elif delta < -max_step:
                delta = -max_step
            self._base_occupancy = max(0.0, self._base_occupancy + delta)

            new_base = int(round(self._base_occupancy))
            base_delta = new_base - self._base_occupancy_rounded
            if base_delta > 0:
                self._sim_entries_total += base_delta
                self._tick_entries_buffer += base_delta
            elif base_delta < 0:
                exits = abs(base_delta)
                self._sim_exits_total += exits
                self._tick_exits_buffer += exits
            self._base_occupancy_rounded = new_base

            temp_offset = 0.0
            active_impacts: Deque[Dict[str, float]] = deque(maxlen=self._detection_impacts.maxlen)
            for item in self._detection_impacts:
                age = now_mono - float(item.get("ts", now_mono))
                if age >= rollback_sec:
                    continue
                weight = max(0.0, 1.0 - (age / rollback_sec))
                temp_offset += float(item.get("delta", 0.0)) * weight
                active_impacts.append(item)
            self._detection_impacts = active_impacts

            self._effective_occupancy = max(0, new_base + int(round(temp_offset)))
            self._last_profile_timestamp = now_dt.isoformat(timespec="seconds")

    def _target_for_datetime(self, when: datetime) -> Tuple[float, float]:
        weekday = int(when.weekday())
        slot = int(when.hour * 4 + (when.minute // 15))

        with self._lock:
            day_slot = self._profile_by_weekday_slot.get((weekday, slot))
            if day_slot is not None:
                return day_slot
            slot_only = self._profile_by_slot.get(slot)
            if slot_only is not None:
                return slot_only
            return self._global_mean, max(1.0, self._global_mean * 0.25)

    def _load_profile_locked(self):
        self._profile_loaded = False
        self._profile_error = ""
        self._profile_bucket_count = 0
        self._profile_by_weekday_slot = {}
        self._profile_by_slot = {}
        self._global_mean = 0.0

        path = self.excel_path
        if not os.path.isabs(path):
            path = os.path.abspath(path)
        if not os.path.exists(path):
            self._profile_error = f"excel_not_found: {path}"
            return

        try:
            import openpyxl
        except Exception as exc:
            self._profile_error = f"openpyxl_missing: {exc}"
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
        except Exception as exc:
            self._profile_error = f"excel_open_failed: {exc}"
            return

        try:
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            col_map = {str(name or "").strip().lower(): idx for idx, name in enumerate(header)}
            ts_idx = col_map.get("timestamp")
            occ_idx = col_map.get("occupancy")
            if ts_idx is None or occ_idx is None:
                self._profile_error = "excel_columns_missing: timestamp/occupancy"
                wb.close()
                return

            buckets: Dict[Tuple[int, int], List[float]] = defaultdict(list)
            slot_buckets: Dict[int, List[float]] = defaultdict(list)
            all_values: List[float] = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                ts_val = row[ts_idx] if ts_idx < len(row) else None
                occ_val = row[occ_idx] if occ_idx < len(row) else None
                if ts_val is None or occ_val is None:
                    continue
                if not isinstance(ts_val, datetime):
                    continue
                try:
                    occ = max(0.0, float(occ_val))
                except Exception:
                    continue

                weekday = int(ts_val.weekday())
                slot = int(ts_val.hour * 4 + (ts_val.minute // 15))
                buckets[(weekday, slot)].append(occ)
                slot_buckets[slot].append(occ)
                all_values.append(occ)

            wb.close()
            if not all_values:
                self._profile_error = "excel_empty_profile"
                return

            self._profile_by_weekday_slot = {
                key: (
                    float(statistics.fmean(values)),
                    float(statistics.pstdev(values)) if len(values) > 1 else max(1.0, float(values[0]) * 0.15),
                )
                for key, values in buckets.items()
            }
            self._profile_by_slot = {
                key: (
                    float(statistics.fmean(values)),
                    float(statistics.pstdev(values)) if len(values) > 1 else max(1.0, float(values[0]) * 0.15),
                )
                for key, values in slot_buckets.items()
            }
            self._global_mean = float(statistics.fmean(all_values))
            try:
                q95 = statistics.quantiles(all_values, n=20, method="inclusive")[18]
                self._max_reasonable = max(10.0, float(q95) * 1.35)
            except Exception:
                self._max_reasonable = max(10.0, max(all_values) * 1.2)
            self._base_occupancy = max(0.0, self._global_mean)
            self._base_occupancy_rounded = int(round(self._base_occupancy))
            self._effective_occupancy = int(self._base_occupancy_rounded)
            self._profile_bucket_count = len(self._profile_by_weekday_slot)
            self._profile_loaded = True
            self._profile_error = ""
        except Exception as exc:
            self._profile_error = f"excel_parse_failed: {exc}"
